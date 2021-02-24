from itertools import combinations
from openpyxl import load_workbook  

chillersToUse = [1500, 1500, 2000, 2000, 3000, 3000, 4000, 4000]

def getAllChillerCombination(chillersToUse, noOfChillersInCombination):
    noOfChillers = len(chillersToUse)
    allChillerCombinations = []
    for i in range(1, noOfChillers + 1):
        allChillerCombinations += (combinations(chillersToUse, i))
    return allChillerCombinations



def read_data_from_file():
    filename = "Combined_data_of_UTown_and_EA_Gokul.xlsx"
    workbook = load_workbook(filename = filename)
    sheet = workbook.active
    print (sheet["R"] [17].value)

    sheet["R10"].value = "1 X 500"
    workbook.save(filename = filename)

##read_data_from_file()

def format_chiller(list_of_chillers):
    if list_of_chillers == "fail":
        return "fail"
    chillers_in_string_format = ""
    while list_of_chillers:
        chiller = list_of_chillers[0]
        chillers_in_string_format += str(list_of_chillers.count(chiller)) + " x " + str(chiller) + " "
        while list_of_chillers.count(chiller):
            list_of_chillers.remove(chiller)
    return chillers_in_string_format
            

def get_chiller_combination(cooling_load, min_part_load, max_part_load):
    all_chiller_combinations = getAllChillerCombination(chillersToUse, len(chillersToUse))
    for combi in all_chiller_combinations:
        capacity = sum(list(combi))
        actual_part_load = (cooling_load / capacity)
##        print ("actual part load: " , actual_part_load)
        if actual_part_load >= 1:
            continue
        if actual_part_load >= min_part_load and actual_part_load <= max_part_load:
##            print (list(combi))
            return list(combi)
    return "fail"

def allocate_chillers():
    filename = "Combined_data_of_UTown_and_EA_Gokul.xlsx"
    workbook = load_workbook(filename = filename)
    sheet = workbook.active
    for row_number in range (17, 234):
        cooling_load = float (sheet["K"][row_number].value) + 50.0
        print ("cooling load: ", cooling_load)
        combi = get_chiller_combination(cooling_load, 0.7, 0.8)
        sheet["R" + str(row_number)].value = format_chiller(combi)
        if type(combi) == list and combi:
            print(combi, row_number)
            sheet["S" + str(row_number)].value = cooling_load/(sum(combi))
        print ("\n")
    workbook.save(filename = filename)


## Main function call
allocate_chillers()

##Testing functions below
        
##combi = get_chiller_combination(950, 0.7, 0.8)
##print (combi)
##print (format_chiller(combi))

##result = getAllChillerCombination(chillersToUse, len(chillersToUse))
##
##print (result)







    

